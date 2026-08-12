import csv
import hashlib
import math
import warnings
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory

import pypdfium2 as pdfium
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps, UnidentifiedImageError

from app.services.object_storage import ObjectStorage

_CHUNK_SIZE = 1024 * 1024
_MAX_NATIVE_TEXT_CHARACTERS_PER_PAGE = 100_000
_SPREADSHEET_ROWS_PER_PAGE = 36
_SPREADSHEET_COLUMNS_PER_PAGE = 8
_SPREADSHEET_MAX_SOURCE_ROWS = 10_000
_SPREADSHEET_MAX_SOURCE_COLUMNS = 64
_SPREADSHEET_MAX_CELL_CHARACTERS = 4_000
_SPREADSHEET_IMAGE_WIDTH = 1_600
_SPREADSHEET_ROW_HEIGHT = 38
_SPREADSHEET_HEADER_HEIGHT = 82


class PreprocessingError(Exception):
    def __init__(self, code: str, safe_message: str) -> None:
        super().__init__(safe_message)
        self.code = code
        self.safe_message = safe_message


@dataclass
class PreprocessedDocument:
    page_paths: tuple[Path, ...]
    page_text: tuple[str | None, ...]
    page_count: int
    _temporary_directory: TemporaryDirectory[str]

    def close(self) -> None:
        self._temporary_directory.cleanup()

    def __enter__(self) -> "PreprocessedDocument":
        return self

    def __exit__(self, *_: object) -> None:
        self.close()


def _download_source(
    storage: ObjectStorage,
    object_key: str,
    destination: Path,
    expected_byte_size: int,
    expected_sha256: str,
    expected_mime_type: str,
) -> None:
    stored_object = storage.open_object(object_key)
    streamed_size = 0
    digest = hashlib.sha256()
    try:
        if stored_object.content_length != expected_byte_size:
            raise PreprocessingError(
                "SOURCE_SIZE_MISMATCH",
                "Stored original size does not match verified metadata",
            )
        if stored_object.content_type != expected_mime_type:
            raise PreprocessingError(
                "SOURCE_TYPE_MISMATCH",
                "Stored original type does not match verified metadata",
            )
        with destination.open("wb") as destination_stream:
            while chunk := stored_object.body.read(_CHUNK_SIZE):
                streamed_size += len(chunk)
                if streamed_size > expected_byte_size:
                    raise PreprocessingError(
                        "SOURCE_SIZE_MISMATCH",
                        "Stored original exceeds verified metadata",
                    )
                digest.update(chunk)
                destination_stream.write(chunk)
    finally:
        stored_object.body.close()

    if streamed_size != expected_byte_size:
        raise PreprocessingError(
            "SOURCE_SIZE_MISMATCH",
            "Stored original size does not match verified metadata",
        )
    if digest.hexdigest() != expected_sha256:
        raise PreprocessingError(
            "SOURCE_HASH_MISMATCH",
            "Stored original does not match its trusted hash",
        )


def _render_pdf(
    source_path: Path,
    output_directory: Path,
    maximum_pages: int,
    render_dpi: int,
    maximum_pixels: int,
    maximum_total_pixels: int,
) -> tuple[tuple[Path, ...], tuple[str | None, ...]]:
    document = pdfium.PdfDocument(source_path)
    try:
        page_count = len(document)
        if page_count < 1:
            raise PreprocessingError("EMPTY_DOCUMENT", "PDF contains no pages")
        if page_count > maximum_pages:
            raise PreprocessingError("PAGE_LIMIT_EXCEEDED", "PDF exceeds the page limit")

        scale = render_dpi / 72
        page_paths: list[Path] = []
        page_text: list[str | None] = []
        total_pixels = 0
        for page_number in range(page_count):
            page = document.get_page(page_number)
            try:
                width, height = page.get_size()
                width_pixels = math.ceil(width * scale)
                height_pixels = math.ceil(height * scale)
                if width_pixels <= 0 or height_pixels <= 0:
                    raise PreprocessingError("INVALID_PAGE", "PDF contains an invalid page")
                if width_pixels * height_pixels > maximum_pixels:
                    raise PreprocessingError(
                        "PIXEL_LIMIT_EXCEEDED",
                        "Rendered PDF page exceeds the pixel limit",
                    )
                total_pixels += width_pixels * height_pixels
                if total_pixels > maximum_total_pixels:
                    raise PreprocessingError(
                        "TOTAL_PIXEL_LIMIT_EXCEEDED",
                        "Rendered PDF exceeds the total pixel limit",
                    )

                bitmap = page.render(scale=scale)
                try:
                    image = bitmap.to_pil()
                    try:
                        output_path = output_directory / f"page-{page_number + 1:04d}.png"
                        rgb = image.convert("RGB")
                        try:
                            rgb.save(output_path, format="PNG")
                        finally:
                            rgb.close()
                        page_paths.append(output_path)
                    finally:
                        image.close()
                finally:
                    bitmap.close()
                text_page = page.get_textpage()
                try:
                    text = text_page.get_text_range().strip()
                    page_text.append(
                        text[:_MAX_NATIVE_TEXT_CHARACTERS_PER_PAGE] if text else None
                    )
                finally:
                    text_page.close()
            finally:
                page.close()
        return tuple(page_paths), tuple(page_text)
    finally:
        document.close()


def _normalize_image(
    source_path: Path,
    output_directory: Path,
    expected_mime_type: str,
    maximum_pixels: int,
) -> tuple[Path, ...]:
    expected_format = {"image/jpeg": "JPEG", "image/png": "PNG"}[expected_mime_type]
    with warnings.catch_warnings():
        warnings.simplefilter("error", Image.DecompressionBombWarning)
        with Image.open(source_path) as source_image:
            if source_image.format != expected_format:
                raise PreprocessingError(
                    "SOURCE_TYPE_MISMATCH",
                    "Stored original does not match verified file type",
                )
            if source_image.width * source_image.height > maximum_pixels:
                raise PreprocessingError(
                    "PIXEL_LIMIT_EXCEEDED",
                    "Image exceeds the pixel limit",
                )
            source_image.load()
            normalized = ImageOps.exif_transpose(source_image)
            try:
                output_path = output_directory / "page-0001.png"
                if "A" in normalized.getbands():
                    rgba = normalized.convert("RGBA")
                    try:
                        rgb = Image.new("RGB", rgba.size, "white")
                        try:
                            rgb.paste(rgba, mask=rgba.getchannel("A"))
                            rgb.save(output_path, format="PNG")
                        finally:
                            rgb.close()
                    finally:
                        rgba.close()
                else:
                    rgb = normalized.convert("RGB")
                    try:
                        rgb.save(output_path, format="PNG")
                    finally:
                        rgb.close()
            finally:
                if normalized is not source_image:
                    normalized.close()
    return (output_path,)


def _spreadsheet_cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if len(text) > _SPREADSHEET_MAX_CELL_CHARACTERS:
        raise PreprocessingError(
            "CELL_LIMIT_EXCEEDED",
            "Spreadsheet contains a cell that exceeds the text limit",
        )
    return text


def _render_spreadsheet_page(
    rows: list[tuple[int, list[str]]],
    source_name: str,
    column_offset: int,
    output_path: Path,
    maximum_pixels: int,
) -> str:
    column_count = min(
        _SPREADSHEET_COLUMNS_PER_PAGE,
        max(len(row) - column_offset for _, row in rows),
    )
    height = _SPREADSHEET_HEADER_HEIGHT + len(rows) * _SPREADSHEET_ROW_HEIGHT
    if _SPREADSHEET_IMAGE_WIDTH * height > maximum_pixels:
        raise PreprocessingError(
            "PIXEL_LIMIT_EXCEEDED",
            "Rendered spreadsheet page exceeds the pixel limit",
        )

    image = Image.new("RGB", (_SPREADSHEET_IMAGE_WIDTH, height), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default(size=18)
    title_font = ImageFont.load_default(size=20)
    row_label_width = 72
    cell_width = (_SPREADSHEET_IMAGE_WIDTH - row_label_width) // column_count
    column_end = column_offset + column_count
    draw.text(
        (12, 10),
        f"{source_name} · columns {column_offset + 1}–{column_end}",
        fill="black",
        font=title_font,
    )
    draw.line(
        (
            0,
            _SPREADSHEET_HEADER_HEIGHT - 1,
            _SPREADSHEET_IMAGE_WIDTH,
            _SPREADSHEET_HEADER_HEIGHT - 1,
        ),
        fill="#b7b7b7",
    )
    for column_index in range(column_count + 1):
        x = row_label_width + column_index * cell_width
        draw.line((x, 48, x, height), fill="#d5d5d5")
        if column_index < column_count:
            draw.text(
                (x + 6, 52),
                f"Column {column_offset + column_index + 1}",
                fill="#444444",
                font=font,
            )

    native_lines = [
        f"Source: {source_name}",
        f"Columns: {column_offset + 1}-{column_end}",
    ]
    for display_index, (source_row_number, row) in enumerate(rows):
        y = _SPREADSHEET_HEADER_HEIGHT + display_index * _SPREADSHEET_ROW_HEIGHT
        draw.line((0, y, _SPREADSHEET_IMAGE_WIDTH, y), fill="#e2e2e2")
        draw.text((8, y + 9), str(source_row_number), fill="#666666", font=font)
        values = row[column_offset:column_end]
        native_lines.append(f"Row {source_row_number}\t" + "\t".join(values))
        for column_index, value in enumerate(values):
            visual_value = value.replace("\r", " ").replace("\n", " ↵ ")
            if len(visual_value) > 28:
                visual_value = f"{visual_value[:27]}…"
            draw.text(
                (
                    row_label_width + column_index * cell_width + 6,
                    y + 9,
                ),
                visual_value,
                fill="black",
                font=font,
            )
    image.save(output_path, format="PNG")
    image.close()
    return "\n".join(native_lines)[:_MAX_NATIVE_TEXT_CHARACTERS_PER_PAGE]


def _render_spreadsheet_rows(
    sources: list[tuple[str, object]],
    output_directory: Path,
    maximum_pages: int,
    maximum_pixels: int,
    maximum_total_pixels: int,
) -> tuple[tuple[Path, ...], tuple[str | None, ...]]:
    page_paths: list[Path] = []
    page_text: list[str | None] = []
    total_pixels = 0
    saw_content = False

    def render_buffer(source_name: str, rows: list[tuple[int, list[str]]]) -> None:
        nonlocal total_pixels
        maximum_columns = max(len(row) for _, row in rows)
        for column_offset in range(0, maximum_columns, _SPREADSHEET_COLUMNS_PER_PAGE):
            if len(page_paths) >= maximum_pages:
                raise PreprocessingError(
                    "PAGE_LIMIT_EXCEEDED",
                    "Spreadsheet exceeds the rendered page limit",
                )
            output_path = output_directory / f"page-{len(page_paths) + 1:04d}.png"
            native_text = _render_spreadsheet_page(
                rows,
                source_name,
                column_offset,
                output_path,
                maximum_pixels,
            )
            with Image.open(output_path) as rendered:
                total_pixels += rendered.width * rendered.height
            if total_pixels > maximum_total_pixels:
                raise PreprocessingError(
                    "TOTAL_PIXEL_LIMIT_EXCEEDED",
                    "Rendered spreadsheet exceeds the total pixel limit",
                )
            page_paths.append(output_path)
            page_text.append(native_text)

    for source_name, source_rows in sources:
        buffer: list[tuple[int, list[str]]] = []
        for source_row_number, raw_row in enumerate(source_rows, start=1):
            if source_row_number > _SPREADSHEET_MAX_SOURCE_ROWS:
                raise PreprocessingError(
                    "ROW_LIMIT_EXCEEDED",
                    "Spreadsheet exceeds the source row limit",
                )
            values = [_spreadsheet_cell_text(value) for value in raw_row]
            while values and not values[-1]:
                values.pop()
            if not values:
                continue
            if len(values) > _SPREADSHEET_MAX_SOURCE_COLUMNS:
                raise PreprocessingError(
                    "COLUMN_LIMIT_EXCEEDED",
                    "Spreadsheet exceeds the source column limit",
                )
            saw_content = True
            buffer.append((source_row_number, values))
            if len(buffer) == _SPREADSHEET_ROWS_PER_PAGE:
                render_buffer(source_name, buffer)
                buffer = []
        if buffer:
            render_buffer(source_name, buffer)

    if not saw_content:
        raise PreprocessingError("EMPTY_DOCUMENT", "Spreadsheet contains no values")
    return tuple(page_paths), tuple(page_text)


def _render_csv(
    source_path: Path,
    output_directory: Path,
    maximum_pages: int,
    maximum_pixels: int,
    maximum_total_pixels: int,
) -> tuple[tuple[Path, ...], tuple[str | None, ...]]:
    raw_sample = source_path.read_bytes()[: 64 * 1024]
    encoding = "utf-8-sig"
    try:
        decoded_sample = raw_sample.decode(encoding)
    except UnicodeDecodeError:
        encoding = "cp1252"
        decoded_sample = raw_sample.decode(encoding)
    try:
        dialect = csv.Sniffer().sniff(decoded_sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    try:
        with source_path.open("r", encoding=encoding, newline="") as stream:
            return _render_spreadsheet_rows(
                [("CSV", csv.reader(stream, dialect))],
                output_directory,
                maximum_pages,
                maximum_pixels,
                maximum_total_pixels,
            )
    except (csv.Error, UnicodeDecodeError) as exc:
        raise PreprocessingError("INVALID_DOCUMENT", "CSV could not be parsed") from exc


def _render_xlsx(
    source_path: Path,
    output_directory: Path,
    maximum_pages: int,
    maximum_pixels: int,
    maximum_total_pixels: int,
) -> tuple[tuple[Path, ...], tuple[str | None, ...]]:
    try:
        workbook = load_workbook(
            source_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            sources = [
                (
                    f"Worksheet: {worksheet.title}",
                    worksheet.iter_rows(values_only=True),
                )
                for worksheet in workbook.worksheets
                if worksheet.sheet_state == "visible"
            ]
            return _render_spreadsheet_rows(
                sources,
                output_directory,
                maximum_pages,
                maximum_pixels,
                maximum_total_pixels,
            )
        finally:
            workbook.close()
    except PreprocessingError:
        raise
    except Exception as exc:
        raise PreprocessingError("INVALID_DOCUMENT", "XLSX could not be parsed") from exc


def preprocess_document(
    storage: ObjectStorage,
    object_key: str,
    expected_byte_size: int,
    expected_sha256: str,
    expected_mime_type: str,
    maximum_pages: int,
    render_dpi: int,
    maximum_pixels: int,
    maximum_total_pixels: int,
) -> PreprocessedDocument:
    temporary_directory = TemporaryDirectory(prefix="cassist-preprocess-")
    directory = Path(temporary_directory.name)
    extension = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "text/csv": ".csv",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    }[expected_mime_type]
    source_path = directory / f"source{extension}"

    try:
        _download_source(
            storage,
            object_key,
            source_path,
            expected_byte_size,
            expected_sha256,
            expected_mime_type,
        )
        if expected_mime_type == "application/pdf":
            page_paths, page_text = _render_pdf(
                source_path,
                directory,
                maximum_pages,
                render_dpi,
                maximum_pixels,
                maximum_total_pixels,
            )
        elif expected_mime_type in {"image/jpeg", "image/png"}:
            page_paths = _normalize_image(
                source_path,
                directory,
                expected_mime_type,
                maximum_pixels,
            )
            page_text = (None,)
        elif expected_mime_type == "text/csv":
            page_paths, page_text = _render_csv(
                source_path,
                directory,
                maximum_pages,
                maximum_pixels,
                maximum_total_pixels,
            )
        else:
            page_paths, page_text = _render_xlsx(
                source_path,
                directory,
                maximum_pages,
                maximum_pixels,
                maximum_total_pixels,
            )
        source_path.unlink()
        return PreprocessedDocument(
            page_paths=page_paths,
            page_text=page_text,
            page_count=len(page_paths),
            _temporary_directory=temporary_directory,
        )
    except PreprocessingError:
        temporary_directory.cleanup()
        raise
    except (
        Image.DecompressionBombError,
        Image.DecompressionBombWarning,
        UnidentifiedImageError,
        pdfium.PdfiumError,
        OSError,
    ) as exc:
        temporary_directory.cleanup()
        raise PreprocessingError("INVALID_DOCUMENT", "Document could not be preprocessed") from exc
    except Exception:
        temporary_directory.cleanup()
        raise
