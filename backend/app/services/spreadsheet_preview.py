import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.services.object_storage import ObjectStorage

PREVIEW_MAX_SHEETS = 5
PREVIEW_MAX_ROWS = 100
PREVIEW_MAX_COLUMNS = 30
PREVIEW_MAX_CELL_CHARACTERS = 500


class SpreadsheetPreviewError(Exception):
    pass


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if len(text) > PREVIEW_MAX_CELL_CHARACTERS:
        return f"{text[:PREVIEW_MAX_CELL_CHARACTERS]}…"
    return text


def _bounded_rows(rows: object) -> tuple[list[list[str]], bool]:
    preview: list[list[str]] = []
    truncated = False
    for row_number, raw_row in enumerate(rows):
        if row_number >= PREVIEW_MAX_ROWS:
            truncated = True
            break
        raw_values = list(raw_row)
        values = [_cell_text(value) for value in raw_values[:PREVIEW_MAX_COLUMNS]]
        while values and not values[-1]:
            values.pop()
        preview.append(values)
        if len(raw_values) > PREVIEW_MAX_COLUMNS:
            truncated = True
    return preview, truncated


def _csv_preview(content: bytes) -> tuple[list[dict[str, object]], bool]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise SpreadsheetPreviewError("CSV could not be decoded") from exc
    sample = text[: 64 * 1024]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    try:
        rows, truncated = _bounded_rows(csv.reader(StringIO(text), dialect))
    except csv.Error as exc:
        raise SpreadsheetPreviewError("CSV could not be parsed") from exc
    return [{"name": "CSV", "rows": rows}], truncated


def _xlsx_preview(content: bytes) -> tuple[list[dict[str, object]], bool]:
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise SpreadsheetPreviewError("XLSX could not be parsed") from exc
    try:
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        truncated = len(visible_sheets) > PREVIEW_MAX_SHEETS
        sheets: list[dict[str, object]] = []
        for sheet in visible_sheets[:PREVIEW_MAX_SHEETS]:
            rows, rows_truncated = _bounded_rows(sheet.iter_rows(values_only=True))
            truncated = truncated or rows_truncated
            sheets.append({"name": sheet.title, "rows": rows})
        return sheets, truncated
    finally:
        workbook.close()


def create_spreadsheet_preview(
    storage: ObjectStorage,
    object_key: str,
    mime_type: str,
) -> tuple[list[dict[str, object]], bool]:
    stored = storage.open_object(object_key)
    try:
        content = stored.body.read()
    finally:
        stored.body.close()
    if mime_type == "text/csv":
        return _csv_preview(content)
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _xlsx_preview(content)
    raise SpreadsheetPreviewError("Document is not a supported spreadsheet")
